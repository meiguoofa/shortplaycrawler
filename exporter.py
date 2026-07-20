import csv
import os

from openpyxl import Workbook

from config import EXPORT_DIR
from models import DailyNewDrama, DramaScreenshot, EpisodeAsset, TranslationJob, get_session


HEADERS = [
    "原剧名", "原简介", "作者", "分类", "原海报URL",
    "翻译语言", "翻译后剧名", "翻译后简介", "新海报URL",
    "剧集数", "剧集URL列表",
    "前三集截图URL",   # 6 个 URL 以 `;` 分隔
    "截图描述",        # 6 条描述以 `;` 分隔
]


def _screenshots_for_job(job: TranslationJob) -> list[DramaScreenshot]:
    """按 (episode_no, position_index) 顺序取该 job 的所有截图行。"""
    db = get_session()
    try:
        return (
            db.query(DramaScreenshot)
            .filter_by(translation_job_id=job.id)
            .order_by(DramaScreenshot.episode_no, DramaScreenshot.position_index)
            .all()
        )
    finally:
        db.close()


def _job_to_row(job: TranslationJob, drama) -> list:
    db = get_session()
    try:
        episodes = db.query(EpisodeAsset).filter_by(
            daily_new_drama_id=drama.id
        ).filter(EpisodeAsset.status == "uploaded").order_by(EpisodeAsset.episode_no).all()
        shots = (
            db.query(DramaScreenshot)
            .filter_by(translation_job_id=job.id)
            .order_by(DramaScreenshot.episode_no, DramaScreenshot.position_index)
            .all()
        )
    finally:
        db.close()

    url_list = "\n".join(ep.object_url for ep in episodes if ep.object_url)
    shot_urls = ";\n".join(s.object_url for s in shots if s.object_url)
    shot_descs = ";\n".join(s.description for s in shots if s.description)
    return [
        drama.title,
        drama.description or "",
        drama.author or "",
        drama.category or "",
        drama.cover_url or "",
        job.target_lang,
        job.translated_title or "",
        job.translated_desc or "",
        job.poster_object_url or "",
        len(episodes),
        url_list,
        shot_urls,
        shot_descs,
    ]


def _build_rows(jobs: list[TranslationJob]) -> list[list]:
    db = get_session()
    try:
        rows = []
        for job in jobs:
            drama = db.query(DailyNewDrama).filter_by(id=job.daily_new_drama_id).first()
            if not drama:
                continue
            rows.append(_job_to_row(job, drama))
        return rows
    finally:
        db.close()


def export_jobs(jobs: list[TranslationJob], fmt: str, out_path: str | None = None) -> str:
    """Export jobs to CSV or XLSX. Returns the absolute path of the written file."""
    rows = _build_rows(jobs)
    headers = list(HEADERS)

    os.makedirs(EXPORT_DIR, exist_ok=True)
    if out_path is None:
        from datetime import datetime
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join(EXPORT_DIR, f"export_{ts}.{fmt}")

    if fmt == "csv":
        # csv.writer auto-quotes cells containing newlines - Excel reads them correctly
        with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(headers)
            for r in rows:
                w.writerow(r)
    elif fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for r in rows:
            ws.append(r)
        # Enable wrap-text on URL-list + screenshot columns for readability
        from openpyxl.styles import Alignment
        for col_idx in (len(headers) - 2, len(headers) - 1, len(headers)):  # 剧集URL / 截图URL / 描述
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        # Sensible column widths
        ws.column_dimensions["A"].width = 30   # 原剧名
        ws.column_dimensions["B"].width = 50   # 原简介
        ws.column_dimensions["E"].width = 40   # 原海报URL
        ws.column_dimensions["H"].width = 50   # 翻译后简介
        ws.column_dimensions["I"].width = 40   # 新海报URL
        ws.column_dimensions["K"].width = 80   # 剧集URL列表
        ws.column_dimensions["L"].width = 60   # 前三集截图URL
        ws.column_dimensions["M"].width = 60   # 截图描述
        wb.save(out_path)
    else:
        raise ValueError(f"unsupported fmt: {fmt}")

    return os.path.abspath(out_path)


def export_batch(batch_id: str, fmt: str, out_path: str | None = None) -> str:
    """Export all TranslationJobs sharing a batch_id."""
    db = get_session()
    try:
        jobs = db.query(TranslationJob).filter_by(batch_id=batch_id).order_by(TranslationJob.id).all()
        if not jobs:
            raise ValueError(f"no jobs found for batch_id={batch_id}")
        return export_jobs(jobs, fmt, out_path=out_path)
    finally:
        db.close()


def export_screenshots_batch(batch_id: str, out_path: str | None = None) -> str:
    """导出批次的截图+描述 Excel（3 列：剧名 / 截图URL(;…) / 描述(;…)）。"""
    os.makedirs(EXPORT_DIR, exist_ok=True)
    if out_path is None:
        from datetime import datetime
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join(EXPORT_DIR, f"screenshots_{ts}.xlsx")

    db = get_session()
    try:
        jobs = db.query(TranslationJob).filter_by(batch_id=batch_id).order_by(TranslationJob.id).all()
        if not jobs:
            raise ValueError(f"no jobs found for batch_id={batch_id}")

        wb = Workbook()
        ws = wb.active
        ws.append(["剧名", "截图URL", "截图描述"])
        for job in jobs:
            drama = db.query(DailyNewDrama).filter_by(id=job.daily_new_drama_id).first()
            if not drama:
                continue
            shots = (
                db.query(DramaScreenshot)
                .filter_by(translation_job_id=job.id)
                .order_by(DramaScreenshot.episode_no, DramaScreenshot.position_index)
                .all()
            )
            shot_urls = ";\n".join(s.object_url for s in shots if s.object_url)
            shot_descs = ";\n".join(s.description for s in shots if s.description)
            ws.append([drama.title, shot_urls, shot_descs])

        from openpyxl.styles import Alignment
        for col_idx in (2, 3):
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 80
        ws.column_dimensions["C"].width = 80
        wb.save(out_path)
        return os.path.abspath(out_path)
    finally:
        db.close()
